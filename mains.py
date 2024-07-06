from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import pandas as pd
import requests
from io import StringIO, BytesIO
from bs4 import BeautifulSoup
import re
import time
import urllib3

# Suppress only the single InsecureRequestWarning from urllib3 needed for this script.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load the Excel file
book = load_workbook('links_of_submission.xlsx')
sheet = book.active

# Load the CSV file and rename a column (this operation doesn't affect the script's main functionality)
submission_format = pd.read_csv('submission_format.csv')
submission_format_2 = submission_format.rename(columns={'h1n1_vaccine':'xyz_vaccine'})

# Setup Selenium WebDriver
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

valid_files_found = False

# Function to safely get attribute and handle stale element exceptions
def safe_get_attribute(element, attribute):
    try:
        return element.get_attribute(attribute)
    except StaleElementReferenceException:
        return None

# Function to modify the GitHub URL to access the raw content
def modify_github_url(url):
    if 'github.com' in url and 'blob' in url:
        url = url.replace('github.com', 'raw.githubusercontent.com').replace('/blob', '')
    return url

# Function to check if two DataFrames are equal
def dataframes_equal(df1, df2):
    try:
        return df1.equals(df2)
    except Exception as e:
        print(f"Error comparing DataFrames: {e}")
        return False

# Function to check CSV submission criteria against two formats
def check_csv_submission_criteria(file_url):
    try:
        # Modify the URL to access the raw content if it's a GitHub link
        file_url = modify_github_url(file_url)
        
        # Create a requests session and disable SSL verification
        session = requests.Session()
        response = session.get(file_url, verify=False)
        response.raise_for_status()

        if 'text/csv' in response.headers.get('Content-Type', '') or file_url.endswith('.csv'):
            csv_content = response.content.decode('utf-8')
        else:
            # Use BeautifulSoup to parse the HTML content
            soup = BeautifulSoup(response.content, 'html.parser')

            # Extract embedded CSV data
            pre_tags = soup.find_all('pre')
            if pre_tags:
                csv_content = pre_tags[0].get_text()
            else:
                print(f"No CSV data found in file {file_url}")
                return False

        # Try to read the CSV data with more robust error handling
        try:
            data = pd.read_csv(StringIO(csv_content), on_bad_lines='skip')

            # Check if the columns and number of rows match any submission format
            if data.shape == submission_format.shape and (not dataframes_equal(data, submission_format) and not dataframes_equal(data, submission_format_2)):
                return True
        except pd.errors.ParserError as e:
            print(f"Parsing error in CSV file {file_url}: {e}")
    except Exception as e:
        print(f"Error checking CSV file {file_url}: {e}")

    return False

# Function to check XLSX submission criteria against two formats
def check_xlsx_submission_criteria(file_url):
    try:
        # Modify the URL to access the raw content if it's a GitHub link
        file_url = modify_github_url(file_url)
        
        # Create a requests session and disable SSL verification
        session = requests.Session()
        response = session.get(file_url, verify=False)
        response.raise_for_status()

        if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('Content-Type', '') or file_url.endswith('.xlsx'):
            xlsx_content = response.content
        else:
            print(f"No XLSX data found in file {file_url}")
            return False

        # Try to read the XLSX data with more robust error handling
        try:
            workbook = load_workbook(filename=BytesIO(xlsx_content))
            sheet = workbook.active
            data = pd.DataFrame(sheet.values)
            
            # Use the first row as column headers
            data.columns = data.iloc[0]
            data = data[1:]

            # Check if the columns and number of rows match any submission format
            if data.shape == submission_format.shape and (not dataframes_equal(data, submission_format) and not dataframes_equal(data, submission_format_2)):
                return True
        except Exception as e:
            print(f"Parsing error in XLSX file {file_url}: {e}")
    except Exception as e:
        print(f"Error checking XLSX file {file_url}: {e}")

    return False

# Function to find .csv or .xlsx files recursively in directories
def find_files_in_directory(url, visited=set()):
    global valid_files_found

    if url in visited:
        return set()
    visited.add(url)

    driver.get(url)
    time.sleep(2)  # Ensure the page is fully loaded
    
    # Find all links on the page
    links = driver.find_elements(By.TAG_NAME, "a")

    # Filter links that end with .csv or .xlsx
    file_urls = set()
    directory_urls = set()

    for link in links:
        href = safe_get_attribute(link, "href")
        if href and href != url:
            if href.endswith(".csv"):
                if check_csv_submission_criteria(href):
                    file_urls.add(href)
                    valid_files_found = True
            elif href.endswith(".xlsx"):
                if check_xlsx_submission_criteria(href):
                    file_urls.add(href)
                    valid_files_found = True
            elif 'tree/' in href:
                directory_urls.add(href)
    
    # If there are directories, search within them recursively
    for directory_url in directory_urls:
        file_urls.update(find_files_in_directory(directory_url, visited))

    return file_urls

# Iterate through each row in the Excel sheet
for row in sheet.iter_rows(values_only=True):  
    url = row[0]
    valid_files_found = False
    message = ""

    if not re.match(r"https://github\.com/.+/.+", url):
        message = f"Invalid GitHub link: + {url}"
    else:
        file_urls = find_files_in_directory(url)
        # Print the URLs of the files
        if file_urls:
            for file_url in file_urls: 
                message += f"Valid files found: {file_url}\n"
        elif not valid_files_found:
            message = f"No files, if present, match the submission criteria in {url}"
    print(message)
    # Add message to the Excel sheet
    sheet.append([url, message]) 

# Save the updated Excel file
book.save('updated_links_of_submission.xlsx')

# Close the WebDriver
driver.quit()